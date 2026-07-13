from pathlib import Path
from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates

from app.auth.dependencies import require_auth, check_module_access
from app.modules.bi.parser import parse_xls, load_saved, clear_saved, list_reports
from app.modules.bi.config import get_bi_config, save_bi_config, recalculate_all_reports

TEMPLATES_DIR = Path(__file__).parent.parent.parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

router = APIRouter(tags=["bi"])


def _backfill_bi(d: dict) -> dict:
    """Adiciona campos novos a relatórios gerados pelo parser antigo."""
    if not d:
        return d

    # Backfill de faturamento em dados sem esses campos
    k = d.get("kpis", {})
    k.setdefault("total_faturamento",  0.0)
    k.setdefault("total_faturado_cnt", 0)
    k.setdefault("pct_faturado",       0.0)
    k.setdefault("a_faturar",          float(k.get("total_producao", 0)))
    k.setdefault("pacientes_com_faltas", 0)
    k.setdefault("total_agendamentos",   k.get("total_agendados", 0))
    k.setdefault("total_confirmados",    0)
    k.setdefault("total_reagendados",    0)
    k.setdefault("total_desmarcados_ag", k.get("total_cancelados", 0))
    k.setdefault("total_faltas_ag",      k.get("total_faltas", 0))

    d.setdefault("perdas_distribuicao",  {})
    d.setdefault("especialidades",       [])
    d.setdefault("faixa_etaria",         {})
    d.setdefault("genero",               {"feminino": 0, "masculino": 0})
    d.setdefault("localidade",           {})
    d.setdefault("convenios_pacientes",  {})
    k.setdefault("idade_media",    0.0)
    k.setdefault("total_acima_60", 0)
    k.setdefault("total_acima_80", 0)
    k.setdefault("pct_acima_60",   0.0)
    k.setdefault("pct_acima_80",   0.0)
    for p in d.get("profissionais", []):
        p.setdefault("desmarcados", p.get("marcado", 0))
        p.setdefault("faltas_ag",   0)
        p.setdefault("reagendados", 0)

    if "alcance_meta" in d:
        return d
    from app.modules.bi.parser import META, META_DIARIA
    k   = d.get("kpis", {})
    prod = float(k.get("total_producao", 0))
    evo = d.get("evolucao_diaria", {})
    n   = len(evo.get("labels", []))
    dias = n or 1

    d["meta"]         = META
    d["meta_diaria"]  = round(META_DIARIA, 2)
    d["alcance_meta"] = round(prod / META * 100, 1)
    d["dias"]         = n
    d["media_diaria"] = round(prod / dias, 2)
    d["melhor_dia"]   = "—"
    d["melhor_dia_prod"] = 0.0
    d.setdefault("cross_data",         {"labels_ag": [], "labels_at": [], "matrix": [], "totals_ag": [], "totals_at": []})
    d.setdefault("status_atendimento", {})
    d.setdefault("atendentes",         d.get("agendadores", []))

    evo.setdefault("prod",     [0.0] * n)
    evo.setdefault("acum",     [0.0] * n)
    evo.setdefault("meta_pct", [0.0] * n)

    for p in d.get("profissionais", []):
        p.setdefault("total",      p.get("qtde", 0))
        p.setdefault("final",      p.get("qtde", 0))
        p.setdefault("marcado",    0)
        p.setdefault("prod",       0.0)
        p.setdefault("taxa_final", 0.0)

    for c in d.get("convenios", []):
        c.setdefault("final", c.get("qtde", 0))

    for t in d.get("tipos", []):
        t.setdefault("final", t.get("qtde", 0))

    for a in d.get("atendentes", []):
        a.setdefault("marcado",    a.get("agendamentos", 0) - a.get("finalizados", 0))
        a.setdefault("taxa_final", round(a.get("finalizados", 0) / max(1, a.get("agendamentos", 1)) * 100, 1))

    return d


@router.get("/bi", response_class=HTMLResponse)
async def bi_dashboard(
    request:     Request,
    period:      str = None,   # compat: mês único antigo
    period_from: str = None,   # compat: AAAA-MM
    period_to:   str = None,   # compat: AAAA-MM
    date_from:   str = None,   # novo: AAAA-MM-DD
    date_to:     str = None,   # novo: AAAA-MM-DD
    unit:        str = None,
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user
    redir = check_module_access(user, "bi")
    if redir:
        return redir

    # Prioridade: date_from/date_to > period_from/period_to > period
    if date_from and not period_from:
        period_from = date_from[:7]
    if date_to and not period_to:
        period_to = date_to[:7]
    if period and not period_from and not period_to:
        period_from = period_to = period

    from app.modules.bi.parser import load_saved_range
    bi_data_raw = _backfill_bi(load_saved_range(period_from, period_to))
    reports     = list_reports()
    bi_cfg      = get_bi_config()

    # Filtro de unidade: substitui os KPIs pelo subconjunto da unidade selecionada
    bi_data       = bi_data_raw
    current_unit  = None
    lista_unidades: list[str] = []

    if bi_data_raw:
        lista_unidades = bi_data_raw.get("lista_unidades", [])

    # Fallback: usa as unidades das metas já salvas para não esconder a seção de parâmetros
    metas_salvas = bi_cfg.get("metas_por_unidade", {})
    for k in metas_salvas:
        if k not in lista_unidades:
            lista_unidades.append(k)
        if unit and unit in bi_data_raw.get("por_unidade", {}):
            unit_data = bi_data_raw["por_unidade"][unit]
            bi_data = {
                **bi_data_raw,
                **unit_data,
                "periodo":          bi_data_raw["periodo"],
                "period_key":       bi_data_raw["period_key"],
                "lista_unidades":   lista_unidades,
                "por_unidade":      {},
                "unidades_summary": bi_data_raw.get("unidades_summary", []),
            }
            current_unit = unit

            # Recalcula profissionais da unidade direto de bi_atendimentos (lista completa, sem cap)
            from app.modules.bi.parser import _calc_profissionais_from_atendimentos
            unit_profs = _calc_profissionais_from_atendimentos(
                period_from or period_to or "", period_to or period_from or "", unit
            )
            if unit_profs:
                bi_data["profissionais"] = unit_profs

            # Aplica meta específica da unidade (se configurada)
            metas_un = bi_cfg.get("metas_por_unidade", {})
            if unit in metas_un and metas_un[unit] > 0:
                meta_u   = float(metas_un[unit])
                dias_u   = float(bi_cfg.get("dias_uteis_mes", 22))
                prod_u   = float(bi_data.get("kpis", {}).get("total_producao", 0))
                meta_dia = meta_u / dias_u if dias_u else meta_u
                bi_data["meta"]         = meta_u
                bi_data["meta_diaria"]  = round(meta_dia, 2)
                bi_data["alcance_meta"] = round(prod_u / meta_u * 100, 1) if meta_u else 0.0
                evo  = bi_data.get("evolucao_diaria", {})
                prods = evo.get("prod", [])
                if prods:
                    evo["meta_pct"] = [round(p / meta_dia * 100, 1) if meta_dia else 0.0 for p in prods]

    # Deriva current_period para compat com listagem/exports que ainda usam period=
    current_period = period_from if period_from == period_to else None

    # Computa date_from/date_to para inputs de data no template
    import calendar as _cal
    if not date_from and period_from:
        date_from = period_from + "-01"
    if not date_to and period_to:
        y, m = int(period_to[:4]), int(period_to[5:7])
        date_to = f"{period_to}-{_cal.monthrange(y, m)[1]:02d}"
    # Fallback: usa o período mais recente disponível
    if not date_from and not date_to and reports:
        latest = reports[0]["period_key"]
        date_from = latest + "-01"
        y, m = int(latest[:4]), int(latest[5:7])
        date_to = f"{latest}-{_cal.monthrange(y, m)[1]:02d}"

    return templates.TemplateResponse("bi.html", {
        "request":         request,
        "user":            user,
        "active_menu":     "bi",
        "bi_data":         bi_data,
        "reports":         reports,
        "current_period":  current_period,
        "period_from":     period_from or "",
        "period_to":       period_to   or "",
        "date_from":       date_from   or "",
        "date_to":         date_to     or "",
        "current_unit":    current_unit,
        "lista_unidades":  lista_unidades,
        "bi_cfg":          bi_cfg,
    })


@router.get("/bi/listagem")
async def bi_listagem(
    request:      Request,
    period:       str  = None,   # compat: mês único
    period_from:  str  = None,   # compat: AAAA-MM
    period_to:    str  = None,   # compat: AAAA-MM
    date_from:    str  = None,   # novo: AAAA-MM-DD
    date_to:      str  = None,   # novo: AAAA-MM-DD
    convenio:     str  = None,
    unidade:      str  = None,
    profissional: str  = None,
    faturado:     str  = None,   # "sim" | "nao" | None
    page:         int  = 1,
    per_page:     int  = 50,
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user
    redir = check_module_access(user, "bi")
    if redir:
        return redir

    # Normaliza period compat
    if period and not period_from:
        period_from = period_to = period

    try:
        from app.database import get_supabase_admin
        sb  = get_supabase_admin()

        def _apply_filters(q, include_faturado: bool = True):
            if date_from:     q = q.gte("data_atend", date_from)
            elif period_from: q = q.gte("period_key", period_from)
            if date_to:       q = q.lte("data_atend", date_to)
            elif period_to:   q = q.lte("period_key", period_to)
            if convenio:     q = q.ilike("convenio", f"%{convenio}%")
            if unidade:      q = q.ilike("unidade", f"%{unidade}%")
            if profissional: q = q.ilike("profissional", f"%{profissional}%")
            if include_faturado:
                if faturado == "sim": q = q.eq("faturado", True)
                elif faturado == "nao": q = q.eq("faturado", False)
            return q

        qb  = _apply_filters(sb.table("bi_atendimentos").select(
            "data_atend,paciente,profissional,especialidade,convenio,unidade,valor,faturado,protocolo_lote,status",
            count="exact"
        ))
        offset = (page - 1) * per_page
        qb = qb.order("data_atend", desc=True).range(offset, offset + per_page - 1)
        res   = qb.execute()
        rows  = res.data or []
        total = res.count or 0

        # Totais sem filtro de faturamento — lote a lote (Supabase limita 1000 linhas/request)
        trows: list[dict] = []
        tot_offset = 0
        while True:
            qbatch = _apply_filters(
                sb.table("bi_atendimentos").select("valor,faturado"),
                include_faturado=False,
            ).eq("status", "realizado").range(tot_offset, tot_offset + 999)
            batch = qbatch.execute().data or []
            trows.extend(batch)
            if len(batch) < 1000:
                break
            tot_offset += 1000
        soma_total    = round(sum(r["valor"] or 0 for r in trows), 2)
        soma_faturado = round(sum(r["valor"] or 0 for r in trows if r["faturado"]), 2)

        return JSONResponse({
            "ok": True, "rows": rows, "total": total,
            "page": page, "per_page": per_page,
            "pages": max(1, -(-total // per_page)),
            "soma_total": soma_total,
            "soma_faturado": soma_faturado,
            "soma_a_faturar": round(soma_total - soma_faturado, 2),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "erro": str(e)}, status_code=500)


@router.get("/bi/listagem/convenios")
async def bi_listagem_convenios(request: Request, period: str = None,
                                 period_from: str = None, period_to: str = None,
                                 date_from: str = None, date_to: str = None,
                                 user=Depends(require_auth)):
    if isinstance(user, RedirectResponse):
        return user
    if period and not period_from:
        period_from = period_to = period
    try:
        from app.database import get_supabase_admin
        sb = get_supabase_admin()
        q  = sb.table("bi_atendimentos").select("convenio")
        if date_from:     q = q.gte("data_atend", date_from)
        elif period_from: q = q.gte("period_key", period_from)
        if date_to:       q = q.lte("data_atend", date_to)
        elif period_to:   q = q.lte("period_key", period_to)
        res = q.limit(5000).execute()
        items = sorted({r["convenio"] for r in (res.data or []) if r.get("convenio")})
        return JSONResponse({"ok": True, "items": items})
    except Exception as e:
        return JSONResponse({"ok": False, "items": [], "erro": str(e)})


@router.get("/bi/listagem/profissionais")
async def bi_listagem_profissionais(request: Request, period: str = None,
                                     period_from: str = None, period_to: str = None,
                                     date_from: str = None, date_to: str = None,
                                     user=Depends(require_auth)):
    if isinstance(user, RedirectResponse):
        return user
    if period and not period_from:
        period_from = period_to = period
    try:
        from app.database import get_supabase_admin
        sb = get_supabase_admin()
        q  = sb.table("bi_atendimentos").select("profissional")
        if date_from:     q = q.gte("data_atend", date_from)
        elif period_from: q = q.gte("period_key", period_from)
        if date_to:       q = q.lte("data_atend", date_to)
        elif period_to:   q = q.lte("period_key", period_to)
        res = q.limit(5000).execute()
        items = sorted({r["profissional"] for r in (res.data or []) if r.get("profissional")})
        return JSONResponse({"ok": True, "items": items})
    except Exception as e:
        return JSONResponse({"ok": False, "items": [], "erro": str(e)})


@router.get("/bi/listagem/export")
async def bi_listagem_export(
    request:      Request,
    period:       str = None,   # compat
    period_from:  str = None,
    period_to:    str = None,
    date_from:    str = None,   # novo: AAAA-MM-DD
    date_to:      str = None,   # novo: AAAA-MM-DD
    convenio:     str = None,
    unidade:      str = None,
    profissional: str = None,
    faturado:     str = None,
    user=Depends(require_auth),
):
    """Exporta a listagem filtrada como Excel (.xlsx)."""
    import io as _io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if isinstance(user, RedirectResponse):
        return user
    redir = check_module_access(user, "bi")
    if redir:
        return redir

    try:
        from app.database import get_supabase_admin
        sb = get_supabase_admin()
        if period and not period_from:
            period_from = period_to = period
        qb = sb.table("bi_atendimentos").select(
            "data_atend,paciente,profissional,especialidade,convenio,unidade,valor,faturado,protocolo_lote,status"
        )
        if date_from:     qb = qb.gte("data_atend", date_from)
        elif period_from: qb = qb.gte("period_key", period_from)
        if date_to:       qb = qb.lte("data_atend", date_to)
        elif period_to:   qb = qb.lte("period_key", period_to)
        if convenio:     qb = qb.ilike("convenio", f"%{convenio}%")
        if unidade:      qb = qb.ilike("unidade",  f"%{unidade}%")
        if profissional: qb = qb.ilike("profissional", f"%{profissional}%")
        if faturado == "sim": qb = qb.eq("faturado", True)
        elif faturado == "nao": qb = qb.eq("faturado", False)
        res  = qb.order("data_atend", desc=True).limit(50000).execute()
        rows = res.data or []

        # ── Monta o workbook ──────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listagem de Produção"

        COLS = [
            ("Data",          "data_atend",     14),
            ("Paciente",      "paciente",        30),
            ("Profissional",  "profissional",    24),
            ("Especialidade", "especialidade",   22),
            ("Convênio",      "convenio",        22),
            ("Unidade",       "unidade",         20),
            ("Valor (R$)",    "valor",           14),
            ("Faturado",      "faturado",        12),
            ("Protocolo Lote","protocolo_lote",  18),
            ("Status",        "status",          14),
        ]

        # Estilo do cabeçalho
        hdr_fill   = PatternFill("solid", fgColor="3D6B1A")
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        hdr_align  = Alignment(horizontal="center", vertical="center")
        thin_side  = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=thin_side, right=thin_side, bottom=thin_side)

        for col_idx, (header, _, width) in enumerate(COLS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.alignment = hdr_align
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Estilo zebra
        fill_even = PatternFill("solid", fgColor="F5FAF2")
        fill_odd  = PatternFill("solid", fgColor="FFFFFF")
        fat_font  = Font(color="1A5C0D", bold=True)
        nfat_font = Font(color="B83030")

        for row_idx, r in enumerate(rows, start=2):
            zebra = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, (_, field, _) in enumerate(COLS, start=1):
                val = r.get(field)
                if field == "data_atend" and val:
                    val = str(val)[:10].split("-")
                    val = f"{val[2]}/{val[1]}/{val[0]}" if len(val) == 3 else r.get(field)
                elif field == "faturado":
                    val = "Sim" if r.get("faturado") else "Não"
                elif field == "valor":
                    try: val = float(val or 0)
                    except: val = 0.0
                elif val is None:
                    val = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill   = zebra
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")

                if field == "valor":
                    cell.number_format = '#,##0.00'
                if field == "faturado":
                    cell.font = fat_font if r.get("faturado") else nfat_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Linha de totais
        tot_row = len(rows) + 2
        ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
        tot_cell = ws.cell(row=tot_row, column=7,
                           value=f"=SUM(G2:G{len(rows)+1})")
        tot_cell.number_format = '#,##0.00'
        tot_cell.font = Font(bold=True, color="1A5C0D")

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"listagem_{period or 'todos'}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        return JSONResponse({"ok": False, "erro": str(e)}, status_code=500)


@router.get("/bi/pacientes", response_class=HTMLResponse)
async def bi_pacientes(
    request:     Request,
    period_from: str = None,
    period_to:   str = None,
    period:      str = None,
    date_from:   str = None,   # AAAA-MM-DD
    date_to:     str = None,
    unit:        str = None,
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user
    redir = check_module_access(user, "bi")
    if redir:
        return redir
    if period and not period_from:
        period_from = period_to = period
    reports = list_reports()
    return templates.TemplateResponse("bi_pacientes.html", {
        "request":      request,
        "user":         user,
        "active_menu":  "bi",
        "reports":      reports,
        "period_from":  period_from,
        "period_to":    period_to,
        "date_from":    date_from,
        "date_to":      date_to,
        "current_unit": unit,
    })


@router.get("/bi/pacientes/data")
async def bi_pacientes_data(
    request:     Request,
    period_from: str = None,   # AAAA-MM (mês inteiro)
    period_to:   str = None,
    period:      str = None,
    date_from:   str = None,   # AAAA-MM-DD (data exata)
    date_to:     str = None,
    unit:        str = None,
    search:      str = None,
    page:        int = 1,
    per_page:    int = 50,
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user
    if period and not period_from:
        period_from = period_to = period
    try:
        from app.database import get_supabase_admin
        sb = get_supabase_admin()

        # Busca todas as linhas do período — tenta incluir tipo_atendimento (pode não existir ainda)
        def _build_query(cols: str):
            q = sb.table("bi_atendimentos").select(cols)
            if date_from:    q = q.gte("data_atend", date_from)   # filtro por data exata
            elif period_from: q = q.gte("period_key", period_from) # fallback por mês
            if date_to:      q = q.lte("data_atend", date_to)
            elif period_to:  q = q.lte("period_key", period_to)
            if unit:   q = q.eq("unidade", unit)
            if search: q = q.ilike("paciente", f"%{search}%")
            return q

        all_rows: list[dict] = []
        has_tipo_col = True
        offset = 0
        while True:
            cols = "paciente,especialidade,tipo_atendimento,status,unidade" if has_tipo_col \
                   else "paciente,especialidade,status,unidade"
            try:
                res = _build_query(cols).range(offset, offset + 999).execute()
            except Exception:
                has_tipo_col = False
                res = _build_query("paciente,especialidade,status,unidade").range(offset, offset + 999).execute()
            batch = res.data or []
            all_rows.extend(batch)
            if len(batch) < 1000:
                break
            offset += 1000

        # Usa tipo_atendimento se disponível, senão especialidade
        def _tipo(r):
            return (r.get("tipo_atendimento") or r.get("especialidade") or "Não informado").strip()

        # Agrega por paciente
        from collections import defaultdict
        pac_map: dict[str, dict] = {}
        for r in all_rows:
            nome = (r.get("paciente") or "").strip()
            if not nome:
                continue
            status = r.get("status") or ""
            tipo   = _tipo(r)
            if nome not in pac_map:
                pac_map[nome] = {"nome": nome, "agendamentos": 0, "finalizados": 0, "tipos": {}}
            pac_map[nome]["agendamentos"] += 1
            if status == "realizado":
                pac_map[nome]["finalizados"] += 1
                pac_map[nome]["tipos"][tipo] = pac_map[nome]["tipos"].get(tipo, 0) + 1

        # Ordena por finalizados desc
        pacs = sorted(pac_map.values(), key=lambda x: -x["finalizados"])
        total = len(pacs)
        total_agend = sum(p["agendamentos"] for p in pacs)
        total_fin   = sum(p["finalizados"]  for p in pacs)
        avg_fin     = round(total_fin / total, 1) if total else 0.0

        # Paginação
        start = (page - 1) * per_page
        page_data = [dict(p) for p in pacs[start:start + per_page]]

        # Serializa tipos como lista ordenada
        for p in page_data:
            p["tipos"] = [{"nome": k, "qtde": v}
                          for k, v in sorted(p["tipos"].items(), key=lambda x: -x[1])]

        return JSONResponse({
            "ok":    True,
            "rows":  page_data,
            "total": total,
            "total_agend": total_agend,
            "total_fin":   total_fin,
            "avg_fin":     avg_fin,
            "page":     page,
            "per_page": per_page,
            "pages":    max(1, -(-total // per_page)),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "erro": str(e)}, status_code=500)


@router.post("/bi/upload")
async def bi_upload(
    request: Request,
    file: UploadFile = File(...),
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user

    role = user.get("role") if isinstance(user, dict) else getattr(user, "role", None)
    if role not in ("admin", "coordenacao"):
        return JSONResponse({"ok": False, "erro": "Sem permissão para importar dados."}, status_code=403)

    if not file.filename.lower().endswith((".xls", ".xlsx", ".html", ".htm")):
        return JSONResponse({"ok": False, "erro": "Formato não suportado. Envie o arquivo .xls exportado pelo sistema."}, status_code=400)

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        return JSONResponse({"ok": False, "erro": "Arquivo muito grande (máx. 50 MB)."}, status_code=400)

    try:
        data = parse_xls(content)
    except ValueError as e:
        return JSONResponse({"ok": False, "erro": str(e)}, status_code=422)
    except Exception as e:
        print("BI UPLOAD ERROR:", repr(e))
        return JSONResponse({"ok": False, "erro": "Erro ao processar o arquivo. Verifique se é o export correto."}, status_code=500)

    k = data["kpis"]
    save_error = data.get("_save_error")
    return JSONResponse({
        "ok":         True,
        "period_key": data["period_key"],
        "periodo":    data["periodo"]["label"],
        "linhas":     data["total_registros"],
        "finalizados": k["total_atendimentos"],
        "producao":   k["total_producao"],
        "save_error": save_error,  # None = salvo, string = erro do Supabase
    })


@router.delete("/bi/reports/{period_key}")
async def bi_delete_report(period_key: str, request: Request, user=Depends(require_auth)):
    if isinstance(user, RedirectResponse):
        return user

    role = user.get("role") if isinstance(user, dict) else getattr(user, "role", None)
    if role not in ("admin", "coordenacao"):
        return JSONResponse({"ok": False, "erro": "Sem permissão."}, status_code=403)

    clear_saved(period_key=period_key)
    return JSONResponse({"ok": True})


@router.post("/bi/metas-unidades")
async def bi_save_metas_unidades(
    request: Request,
    metas_json: str = Form(...),
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user
    role = user.get("role") if isinstance(user, dict) else getattr(user, "role", None)
    if role not in ("admin", "coordenacao"):
        return JSONResponse({"ok": False, "erro": "Sem permissão."}, status_code=403)

    import json as _json
    try:
        metas = _json.loads(metas_json)
        if not isinstance(metas, dict):
            raise ValueError("Esperado objeto JSON")
        metas = {k: float(v) for k, v in metas.items() if v}
    except Exception as e:
        return JSONResponse({"ok": False, "erro": f"Dados inválidos: {e}"}, status_code=422)

    err = save_bi_config({"metas_por_unidade": metas})
    return JSONResponse({"ok": True, "aviso": err})


@router.post("/bi/parametros")
async def bi_save_parametros(
    request: Request,
    meta_mensal:      float = Form(...),
    dias_uteis_mes:   float = Form(...),
    cidade_geocoding: str   = Form(""),
    user=Depends(require_auth),
):
    if isinstance(user, RedirectResponse):
        return user

    role = user.get("role") if isinstance(user, dict) else getattr(user, "role", None)
    if role not in ("admin", "coordenacao"):
        return JSONResponse({"ok": False, "erro": "Sem permissão."}, status_code=403)

    if meta_mensal <= 0 or dias_uteis_mes <= 0:
        return JSONResponse({"ok": False, "erro": "Valores devem ser maiores que zero."}, status_code=422)

    cfg_payload = {"meta_mensal": meta_mensal, "dias_uteis_mes": dias_uteis_mes}
    if cidade_geocoding.strip():
        cfg_payload["cidade_geocoding"] = cidade_geocoding.strip()
    cfg_err = save_bi_config(cfg_payload)

    n_recalc, recalc_err = recalculate_all_reports(meta_mensal, dias_uteis_mes)

    avisos = [e for e in [cfg_err, recalc_err] if e]
    return JSONResponse({
        "ok":          True,
        "recalculados": n_recalc,
        "aviso":       "; ".join(avisos) if avisos else None,
    })
